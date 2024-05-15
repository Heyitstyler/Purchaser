import os
import sys
import csv
import glob
import numpy
import pandas as pd
import time
import selenium
import webdriver_manager
import tkinter as tk
from tkinter import ttk
from tkinter import *
from tkinter import filedialog
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By


def dlItemsList(barSelect):
    global sum_e
    exportJS = 'exportItemsReport()'
    found_Sum = "False"
    try:
        sum_e = "Failed"
        os.chdir(dirDL)
        keyword = 'Items'
        options = Options()
        options.set_preference("browser.download.folderList", 2)
        options.set_preference("browser.download.manager.showWhenStarting", False)
        options.set_preference("browser.download.dir", dirDL)
        options.set_preference("browser.helperApps.neverAsk.saveToDisk", "application/x-gzip")
        options.add_argument("--headless")


        summary_driver = webdriver.Firefox(options=options)
        sumWait = WebDriverWait(summary_driver, 90)

        summary_driver.get("https://www.barkeepapp.com/BarkeepOnline/items.php")

        login_Loaded = sumWait.until(EC.presence_of_element_located((By.NAME, 'session_username')))
        username_field = summary_driver.find_element(By.NAME, 'session_username')
        username_field.send_keys(barSelect)
        password_field = summary_driver.find_element(By.NAME, 'session_password')
        password_field.send_keys(passwd)
        login_button = summary_driver.find_element(By.NAME, 'login')
        login_button.click()

        summary_driver.execute_script(exportJS)

        while found_Sum == "False":
        # List all files in the specified directory
            files = os.listdir(dirDL)

        # Check if any file contains the keyword
            for file in files:
                if file.startswith(keyword) and not file.endswith(".part"):
                    print(f"Found file: {file}")
                    sum_e = (f"{proper} Summary report")
                    time.sleep(1)
                    summary_driver.close()
                    time.sleep(0.5)
                    summary_driver.quit()
                    os.chdir(dirroot)
                    found_Sum = "True"
                    return
        
        
    except:
        sum_e = ("Error Collecting Summary Report")
        summary_driver.close()
        time.sleep(1)
        summary_driver.quit()
        os.chdir(dirroot)
        log = open("dllog.txt", "a")
        L = [f"Failed Summary Report\n"]
        log.writelines(L)
        log.close()
        return


def Items_to_list():
    global unique_data
    barSum = glob.glob(os.path.join(dirDL, "Items*.xlsx"))[0]
    os.chdir(dirDL)
    wb = load_workbook(barSum)
    ws = wb.active

    data = []
    for row in ws['E']:
        if row.value is not None:
            data.append(row.value)
    
    unique_data = list(set(data))
    

def invoiceImport():
    global FinInvoice
    try:
        
        os.chdir(invoicePath)
        invoice = glob.glob(os.path.join(invoicePath, "invoicedetail*.csv"))[0]
        df = pd.read_csv(invoice)
        selected_collumns = ['Product Description', 'Unit Of Measure', 'Quantity', 'Unit Cost']
        trimmedInvoice = df[selected_collumns]
        os.chdir(dirDL)
        trimmedInvoice.to_excel('invoice.xlsx', index=False)
        FinInvoice = glob.glob(os.path.join(dirDL, 'invoice.xlsx'))[0]

    except Exception as e:
        print(f"{e}")


def updateItemDB():
    global currentDB
    os.chdir(barDB)
    DBName = username + 'DB.xlsx'
    newFull = unique_data
    try:
        currentDB = glob.glob(os.path.join(barDB, DBName))[0]
    except:
        pass
    if os.path.exists(DBName):
        wb = load_workbook(currentDB)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['Barkeep', 'Fintech', 'Case'])
        wb.save(DBName)
        currentDB = glob.glob(os.path.join(barDB, DBName))[0]
    
    
    if ws.max_row > 1:  # Check if there is more than just the header row
        df_excel = pd.DataFrame(ws.values)
        existing_items = df_excel[0].dropna().tolist()  # Column C index is 2 (0-based)
    else:
        existing_items = []

    # df_csv = pd.read_csv(itemsCsv, header=None)
    new_items = unique_data

    items_to_add = [item for item in new_items if item not in existing_items]

    for item in items_to_add:
        ws.append([item])

    wb.save(currentDB)


def load_excel_items_with_rows(file_path, columns):
    wb = load_workbook(file_path)
    ws = wb.active
    items = {col: {} for col in columns}
    for col in columns:
        for row, cell in enumerate(ws[col], start=1):
            if cell.value is not None:
                items[col][cell.value] = row
    return items, wb, ws


def update_excel(ws, wb, selected_item, unmatched_item, casesize, file_path):
    os.chdir(barDB)
    row = items['A'].get(selected_item)
    if row:
        # Column B is usually index 2 (as columns are 1-based in openpyxl)
        current_value = ws.cell(row=row, column=2).value
        if current_value:
            current_items = [item.strip() for item in current_value.split(', ')]
            if unmatched_item not in current_items:
                new_value = f"{current_value}, {unmatched_item}"
            else:
                new_value = current_value
        else:
            new_value = unmatched_item
        
        ws.cell(row=row, column=2, value=new_value)  # Update the cell with the new value
        ws.cell(row=row, column=3, value=casesize)  # Update or set the case size in Column C
        wb.save(file_path)  # Save the workbook
        print(f"Matched '{unmatched_item}' to '{selected_item}' in row {row}")
        completedItems.append(unmatched_item)
    else:
        print("Error: Selected item not found in column A.")


def init_skipped_workbook():
    os.chdir(dirDL)
    skipped_wb = Workbook()
    skipped_liquor_ws = skipped_wb.active
    skipped_liquor_ws.title = 'Skipped Liquor'
    skipped_liquor_ws.append(['Skipped Item'])  # Add a header or initial row
    # Create the second sheet named 'Trash'
    trash_ws = skipped_wb.create_sheet('Trash')
    trash_ws.append(['Trash Item'])  # Add a header or initial row
    skipped_wb.save('skippeditems.xlsx')

    return skipped_wb, skipped_liquor_ws, trash_ws

def init_global_trash():
    global Global_Trash
    os.chdir(barDB)
    if not os.path.exists(os.path.join(barDB, "Global_Trash.xlsx")):
        global_trash_wb = Workbook()
        global_trash_ws = global_trash_wb.active
        global_trash_ws.title = "Global Trash"
        global_trash_ws.append(['Global Trash Items'])
        global_trash_wb.save('Global_Trash.xlsx')
        gtrash_File = glob.glob(os.path.join(barDB, "Global_Trash.xlsx"))[0]

    else:
        gtrash_File = glob.glob(os.path.join(barDB, "Global_Trash.xlsx"))[0]
        global_trash_wb = load_workbook(gtrash_File)
        global_trash_ws = global_trash_wb.active

    gTrashdata = []
    for row in global_trash_ws['A']:
        if row.value is not None:
            gTrashdata.append(row.value)
    
    Global_Trash = list(set(gTrashdata))

    return global_trash_ws, global_trash_wb



def write_global_trash(global_trash_ws, global_trash_wb, item, unmatched_items):
    os.chdir(barDB)
    global_trash_ws.append([item])
    global_trash_wb.save('Global_Trash.xlsx')
    if item in unmatched_items:
        unmatched_items.remove(item)
    print (f'Trashed {item} Globally')

def write_skipped_liquor(skipped_wb, skipped_liquor_ws, trash_ws, item, unmatched_items):
    os.chdir(dirDL)
    skipped_liquor_ws.append([item])
    skipped_wb.save('skippeditems.xlsx')
    completedItems.append(item)
    if item in unmatched_items:
        unmatched_items.remove(item)
    print(f"Skipped and recorded: {item}")

def write_trash(skipped_wb, skipped_liquor_ws, trash_ws, item, unmatched_items):
    os.chdir(dirDL)
    trash_ws.append([item])
    skipped_wb.save('skippeditems.xlsx')
    completedItems.append(item)
    if item in unmatched_items:
        unmatched_items.remove(item)
    print(f"Trashed {item}")



def init_tkinter(choices, unmatched_item, ws, wb, file_path, skipped_wb, skipped_ws, unmatched_items):
    def on_button_clicked():
        selected_item = list_Box.get(list_Box.curselection())
        casesize = casebox.get()
        update_excel(ws, wb, selected_item, unmatched_item, casesize, file_path)
        root.destroy()

    def on_skip_liquor():
        write_skipped_liquor(skipped_wb, skipped_ws, trash_ws, unmatched_item, unmatched_items)
        root.destroy()

    def on_trash_item():
        write_trash(skipped_wb, skipped_ws, trash_ws, unmatched_item, unmatched_items)
        root.destroy()

    def on_global_trash():
        write_global_trash(global_trash_ws, global_trash_wb, unmatched_item, unmatched_items)
        root.destroy()

    def conf_global_trash():
        global conf_root
        conf_root = tk.Tk()
        conf_root.title("Confirm Global Trash")

        item_label = Label(conf_root, text=f'are you sure you want to Globaly Trash\n{unmatched_item}?')
        item_label.grid(column=0, row=0, columnspan=2)

        gTrash_conf = Button(conf_root, text="Confirm", command=lambda:[on_global_trash(), conf_root.destroy()])
        gTrash_conf.grid(column=0, row=1)

        gTrash_cancel = Button(conf_root, text='Cancel', command=lambda:conf_root.destroy())
        gTrash_cancel.grid(column=1, row=1)

    def update(event):
        # Clear the current list
        list_Box.delete(0, tk.END)
        # Get the user input
        current_text = entry.get().lower()
        if current_text:
            # Filter the list and update the Listbox
            filtered_choices = [choice for choice in choices if current_text in choice.lower()]
            for item in filtered_choices:
                list_Box.insert(tk.END, item)
        else:
            # If no input, show all choices
            for choice in choices:
                list_Box.insert(tk.END, choice)

    def enButton(*args):
        if(len(casebox.get())>0):
            button_update.config(state='normal')
        else:
            button_update.config(state='disabled')
    

    root = tk.Tk()
    root.title("Match Item")
    label = ttk.Label(root, text=f"Select a match for: \n{unmatched_item}", font=("Times", "12"))
    label.grid(row=0, column=0, columnspan=4, pady=20, padx=20)

    # Create an Entry widget
    entry = tk.Entry(root, width=35)
    entry.grid(row=1, column=0, columnspan=2,pady=10, padx=10)
    entry.bind('<KeyRelease>', update)

    autocomp = tk.StringVar()

    sorted_choices = sorted(choices.keys())

    scrollbar = Scrollbar(root)
    list_Box = Listbox(root, width=35)
    list_Box.grid(row=2, column=0, columnspan=2, rowspan=4, pady=3, padx=10)

    for choice in sorted_choices:
        list_Box.insert(tk.END, choice)


    casebox = StringVar(root)

    case_label = ttk.Label(root, text="Case Size")
    case_label.grid(row=6, column=0, pady=5)

    case_entry = ttk.Entry(root, text="Case Size", textvariable=casebox)
    case_entry.grid(row=6, column=1, pady=10)
    
    button_update = ttk.Button(root, text="Update Excel", command=on_button_clicked, width=15, state='disabled')
    button_update.grid(row=2, column=2, columnspan=2, pady=3, padx=10, sticky='NW')

    liquor_skip = ttk.Button(root, text="Skip Liquor", command=on_skip_liquor, width=15)
    liquor_skip.grid(row=3, column=2, columnspan=2, pady=3, padx=10, sticky='NW')

    trash_skip = ttk.Button(root, text="Trash Item", command=on_trash_item, width=15)
    trash_skip.grid(row=4, column=2, columnspan=2, pady=3, padx=10, sticky='NW')

    global_trash_Button = ttk.Button(root, text="Global Trash", command=conf_global_trash, width=15)
    global_trash_Button.grid(row=5, column=2, columnspan=2, pady=10, padx=10, sticky='NW')

    casebox.trace_add('write', enButton)
    root.mainloop()



def correctInvoice():
    # Load the Excel files
    invoice_df = pd.read_excel(FinInvoice)  # Assuming names to correct are in a column named 'Product Description'
    item_database_df = pd.read_excel(currentDB, header=None)  # Assuming no header

    # Create a dictionary for mapping incorrect names to correct names
    name_mapping = {}
    for index, row in item_database_df.iterrows():
        incorrect_names = str(row[1]).split(', ')  # Split the names based on ", "
        correct_name = row[0]
        for name in incorrect_names:
            name_mapping[name] = correct_name

    # Create a dictionary for mapping correct names to case sizes
    case_size_mapping = dict(zip(item_database_df[0], item_database_df[2]))

    # Replace the incorrect names in the invoice dataframe
    invoice_df['Product Description'] = invoice_df['Product Description'].map(name_mapping).fillna(invoice_df['Product Description'])

    # Add case sizes to the invoice dataframe
    invoice_df['Case Size'] = invoice_df['Product Description'].map(case_size_mapping)

    # Set directory for saving the file (specify the path if needed)
    os.chdir(dirDL)  # Change to your desired directory

    # Save the corrected DataFrame back to an Excel file
    invoice_df.to_excel('corrected_invoice.xlsx', index=False)


def splitInvoice():
    global new_invoice, skipped_items_df

    def prepare_sheet(wb, sheet_name, main_sheet, num_cols):
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            sheet = wb.create_sheet(sheet_name)
            # Copy headers from the first row of the main_sheet to the new sheet
            for col in range(1, num_cols + 1):  # Adjust num_cols to copy the correct number of column headers
                header_value = main_sheet.cell(row=1, column=col).value
                sheet.cell(row=1, column=col).value = header_value
        return sheet


    os.chdir(dirDL)
    skipped_items_df = pd.read_excel('skippeditems.xlsx')
    trash_items_df = pd.read_excel('skippeditems.xlsx', sheet_name="Trash")

    skipped_names = set(skipped_items_df.iloc[:, 0].dropna())  # Assumes names are in the first column
    trash_names = set(trash_items_df.iloc[:, 0].dropna())

    # Load the workbook and select the main worksheet
    os.chdir(dirDL)

    wb = load_workbook('corrected_invoice.xlsx')
    main_sheet = wb.active

# Create a new worksheet for skipped items
    if 'Skipped Items' in wb.sheetnames:
        skipped_sheet = wb['Skipped Items']
    else:
        skipped_sheet = prepare_sheet(wb, 'Skipped Items', main_sheet, 5)


    if "Trash" in wb.sheetnames:
        trash_sheet = wb["Trash"]
    else:
        trash_sheet = prepare_sheet(wb, 'Trash', main_sheet, 5)

    # Collect rows to be moved and their indices
    rows_to_move = []
    for idx, row in enumerate(main_sheet.iter_rows(min_row=2, max_col=1, values_only=True), start=2):  # Adjust max_col as needed
        if row[0] in skipped_names:
            rows_to_move.append(idx)

    # Append rows in reverse order to avoid changing indices
    for idx in reversed(rows_to_move):
        row_data = [cell.value for cell in main_sheet[idx]]
        skipped_sheet.append(row_data)
        main_sheet.delete_rows(idx)

    wb.save('updated_corrected_invoice.xlsx')



    # Collect rows to be moved and their indices
    trash_to_move = []
    for idx2, row in enumerate(main_sheet.iter_rows(min_row=2, max_col=1, values_only=True), start=2):  # Adjust max_col as needed
        if row[0] in trash_names:
            trash_to_move.append(idx2)

    # Append rows in reverse order to avoid changing indices
    for idx2 in reversed(trash_to_move):
        row_data = [cell.value for cell in main_sheet[idx2]]
        trash_sheet.append(row_data)
        main_sheet.delete_rows(idx2)
    # Save the updated workbook
    wb.save('updated_corrected_invoice.xlsx')
    
    new_invoice = glob.glob(os.path.join(dirDL, "updated_corrected_invoice.xlsx"))[0]


def Purchase():
    #javascript

    toAppend = load_workbook(new_invoice)
    missingItems = toAppend['Skipped Items']
    



    startItemJS = "addInventoryItem()"
    defaultPriceJS = "makeDefaultPrice()"
    addContinueJS = 'createAndContinueEII()'
    addExitJS = 'createAndExitEII()'
    try:
        options = Options()
        options.set_preference("browser.download.folderList", 2)
        options.set_preference("browser.download.manager.showWhenStarting", False)
        options.set_preference("browser.download.dir", dirDL)
        options.set_preference("browser.helperApps.neverAsk.saveToDisk", "application/x-gzip")
        # options.add_argument("--headless")

        purchaseDriver = webdriver.Firefox(options=options)
        purchaseWait = WebDriverWait(purchaseDriver, 30)

        purchaseDriver.get("https://www.barkeepapp.com/BarkeepOnline/inventories.php")

        login_Loaded = purchaseWait.until(EC.presence_of_element_located((By.NAME, 'session_username')))
        username_field = purchaseDriver.find_element(By.NAME, 'session_username')
        username_field.send_keys(barSelect)
        password_field = purchaseDriver.find_element(By.NAME, 'session_password')
        password_field.send_keys(passwd)
        login_button = purchaseDriver.find_element(By.NAME, 'login')
        login_button.click()

        inv_loaded = purchaseWait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[4]/div/div[3]/div[2]/table/tbody/tr[2]/td[1]/a[1]')))
        recievingInv = purchaseDriver.find_element(By.XPATH, '/html/body/div/div[4]/div/div[3]/div[2]/table/tbody/tr[2]/td[1]/a[1]')
        recievingInv.click()

        rec_loaded = purchaseWait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[4]/h2/button")))
        purchaseDriver.execute_script(startItemJS)

        int_loaded = purchaseWait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[3]/div[2]/div/div/div[2]/div[2]/table/tbody/tr[1]/td[1]/button")) or EC.presence_of_element_located((By.XPATH, "/html/body/div[3]/div[2]/div/div/div[2]/div[2]/table/tbody/tr/td[1]/button")))

        # Load the Excel file
        df = pd.read_excel(new_invoice)

        # Iterate through each row in the DataFrame
        for index, row in df.iterrows():
            # Assign each value to a variable named after the column
            product_description = row['Product Description']
            unit_of_measure = row['Unit Of Measure']
            quantity = row['Quantity']
            unit_cost = row['Unit Cost']
            case_size = row['Case Size']

            try:
                product_description = product_description.replace("75ml", "750ml")
            except:
                pass

            try:
                product_description = product_description.replace("1,000ml", "1,0")
            except:
                pass

            try:
                product_description = product_description.replace("1,0ml", "1,0")
            except:
                pass

            try:
                product_description = product_description.strip('"')
            except:
                pass
            

            itemSearch = purchaseDriver.find_element(By.XPATH, "/html/body/div[3]/div[2]/div/div/div[1]/div/label/input")
            if product_description in Global_Trash:
                continue
            itemSearch.send_keys(product_description)
            time.sleep(.5)
            try:
                try:
                    selectItem = purchaseDriver.find_element(By.XPATH, "/html/body/div[3]/div[2]/div/div/div[2]/div[2]/table/tbody/tr/td[1]/button")
                except:
                    selectItem = purchaseDriver.find_element(By.XPATH, "/html/body/div[3]/div[2]/div/div/div[2]/div[2]/table/tbody/tr[1]/td[1]/button")
            except:
                
                # write_skipped_liquor(skipped_wb, skipped_ws, trash_ws, product_description, unmatched_items)
                missingItems.append([product_description, unit_of_measure, quantity, unit_cost, case_size])
                toAppend.save('updated_corrected_invoice.xlsx')
                print(f"Skipped {product_description}")
                itemSearch.clear()
                continue
            selectItem.click()
            time.sleep(.5)

            if unit_of_measure == 'CA' or unit_of_measure == 'CS':
                bottles = int(quantity) * int(case_size)
                price = round(float(unit_cost.strip('$').replace(',', '')) / float(case_size), 4)
            else:
                bottles = quantity
                price = round(float(unit_cost.strip('$').replace(',', '')), 4)


            quantWait = purchaseWait.until(EC.presence_of_element_located((By.ID, 'quantity')))
            quant_Box = purchaseDriver.find_element(By.ID, 'quantity')
            time.sleep(.15)
            for character in str(bottles):
                quant_Box.send_keys(character)
                time.sleep(0.15)

            price_Box = purchaseDriver.find_element(By.ID, 'price')
            price_Box_Value = purchaseDriver.find_element(By.ID, 'price').get_attribute("value").replace('$', '')
            if price_Box_Value != price:
                price_Box.clear()
                for character in str(price):
                    price_Box.send_keys(character)
                    time.sleep(0.15)
                time.sleep(.25)
                if price > 0:
                    purchaseDriver.execute_script(defaultPriceJS)
                    time.sleep(.5)

                    defContWait = purchaseWait.until(EC.presence_of_element_located((By.XPATH, "//button[contains(.,'Ok')]")))

                    defCont = purchaseDriver.find_element(By.XPATH, "//button[contains(.,'Ok')]")
                    defCont.click()
                else:
                    pass
            else:
                pass


            time.sleep(.5)

            purchaseDriver.execute_script(addExitJS)
            time.sleep(3)

            purchaseWait.until(EC.invisibility_of_element((By.XPATH, "/html/body/div[5]")))

            time.sleep(1)

            purchaseDriver.execute_script(startItemJS)

            int_loaded = purchaseWait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[3]/div[2]/div/div/div[2]/div[2]/table/tbody/tr[1]/td[1]/button")) or EC.presence_of_element_located((By.XPATH, "/html/body/div[3]/div[2]/div/div/div[2]/div[2]/table/tbody/tr/td[1]/button")))

            # You can now use these variables for your data entry tasks
            print(product_description, unit_of_measure, quantity, unit_cost, case_size)


        purchaseDriver.quit()

    except Exception as e:
        print(e)
        purchaseDriver.quit()

def notify():
    saveJS = 'saveChanges(null)'
    os.chdir(dirroot)
    upInv = glob.glob(os.path.join(dirDL, 'updated_corrected_invoice.xlsx'))[0]
    skipped_items = pd.read_excel(upInv, sheet_name='Skipped Items', usecols=['Product Description', 'Unit Of Measure', 'Quantity', 'Unit Cost'])
    missing_items = skipped_items.to_csv(index=False, header=False)
    try:
        options = Options()
        options.set_preference("browser.download.folderList", 2)
        options.set_preference("browser.download.manager.showWhenStarting", False)
        options.set_preference("browser.download.dir", dirDL)
        options.set_preference("browser.helperApps.neverAsk.saveToDisk", "application/x-gzip")
        options.add_argument("--headless")

        notesDriver = webdriver.Firefox(options=options)
        notesWait = WebDriverWait(notesDriver, 30)

        notesDriver.get("https://www.barkeepapp.com/BarkeepOnline/inventories.php")

        login_Loaded = notesWait.until(EC.presence_of_element_located((By.NAME, 'session_username')))
        username_field = notesDriver.find_element(By.NAME, 'session_username')
        username_field.send_keys(barSelect)
        password_field = notesDriver.find_element(By.NAME, 'session_password')
        password_field.send_keys(passwd)
        login_button = notesDriver.find_element(By.NAME, 'login')
        login_button.click()

        inv_loaded = notesWait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[4]/div/div[3]/div[2]/table/tbody/tr[2]/td[1]/a[1]')))
        recievingInv = notesDriver.find_element(By.XPATH, '/html/body/div/div[4]/div/div[3]/div[2]/table/tbody/tr[2]/td[1]/a[1]')
        recievingInv.click()

        rec_loaded = notesWait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[4]/h2/button")))


        notesBox = notesDriver.find_element(By.ID, "notes")
        time.sleep(.5)
        notesBox.send_keys(missing_items)
        time.sleep(.5)
        notesDriver.execute_script(saveJS)
        time.sleep(2)
        notesDriver.quit()

    except Exception as e:
        print(e)
        notesDriver.quit()


def Reset():
    for item in os.listdir(dirDL):
        item_path = os.path.join(dirDL, item)
        if os.path.isfile(item_path):
            os.remove(item_path)


def ResetInvoice():
    for item in os.listdir(invoicePath):
        item_path = os.path.join(invoicePath, item)
        if os.path.isfile(item_path):
            os.remove(item_path)





def askResetInvoice():
    def ResetInvoice():
        askReset.destroy()
        for item in os.listdir(invoicePath):
            item_path = os.path.join(invoicePath, item)
            if os.path.isfile(item_path):
                os.remove(item_path)
    try:
        askReset = tk.Tk()
        askLabel = Label(askReset, text='Do you want to delete the invoice?')
        askLabel.grid(row=0, column=0, columnspan=2)

        askResetYes = Button(askReset, text='Yes', command=lambda: ResetInvoice())
        askResetYes.grid(row=1, column=0)

        askResetNo = Button(askReset, text='No', command=askReset.destroy)
        askResetNo.grid(row=1, column=1)

        askReset.mainloop()
    except Exception as e:
        print(e)



if __name__ == '__main__':
    dirroot = os.getcwd()
    dirDL = os.path.join(dirroot, "cache")
    barDB = os.path.join(dirroot, "barDB")
    invoicePath = os.path.join(dirroot, "invoice")

    bars = pd.read_csv(os.path.join(barDB, "bardb.csv"))

    completedItems = []

    while True:
        barSelect = input("What bar are we working with: ")

        userRow = bars[bars["user"] == barSelect]

        if userRow.empty:
            print("Username not found. Please try again.")
            continue
        else:
            break
    

    username = userRow["user"].iloc[0]
    passwd = userRow["pass"].iloc[0]
    proper = userRow["proper"].iloc[0]
    street = userRow["street"].iloc[0]
    city = userRow["city"].iloc[0]
    inv = userRow["invoicename"].iloc[0]
    price = userRow["price"].iloc[0]
    finUser = userRow["finUser"].iloc[0]
    finPass = userRow["finPass"].iloc[0]


    invoiceImport()
    dlItemsList(barSelect)
    Items_to_list()

    updateItemDB()

        # Load items from Excel file
    file_path = currentDB
    columns = ['A', 'B']
    items, wb, ws = load_excel_items_with_rows(file_path, columns)
    skipped_wb, skipped_ws, trash_ws = init_skipped_workbook()

    # Load CSV and iterate through "Product Description" column


    global_trash_ws, global_trash_wb = init_global_trash()


    csv_path = FinInvoice
    df_invoice = pd.read_excel(FinInvoice)
    df_items = pd.read_excel(currentDB)
    all_items_b = set()
    df_items['Fintech'].dropna().apply(lambda x: all_items_b.update(x.split(', ')))
    unmatched_items = [item for item in df_invoice['Product Description'] if item not in all_items_b and Global_Trash]

    for unmatched_item in unmatched_items[:]:  # Use a slice copy to iterate while potentially modifying the list
        if unmatched_item not in Global_Trash[:]:
            if unmatched_item not in completedItems[:]:
                init_tkinter(items['A'], unmatched_item, ws, wb, file_path, skipped_wb, skipped_ws, unmatched_items)


    correctInvoice()
    splitInvoice()
    Purchase()
    
    notify()

    Reset()
    askResetInvoice()